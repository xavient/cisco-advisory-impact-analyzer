"""
Excel report writer.

Produces analysis_output_<timestamp>.xlsx with three columns matching the team's
reference format: Vendor Advisory#, Effected Product Description, Expected Assessment.
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = ["Vendor Advisory#", "Effected Product Description", "Expected Assessment"]


def write_report(results, output_dir="."):
    """Write results (list of (advisory_id, description, assessment)) to a new xlsx.

    Returns the path written.
    """
    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(output_dir) / f"analysis_output_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for advisory_id, description, assessment in results:
        ws.append([advisory_id, description, assessment])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 30
    ws.freeze_panes = "A2"

    wb.save(out_path)
    return str(out_path)
