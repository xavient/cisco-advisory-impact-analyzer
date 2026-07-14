"""
Firewall inventory loading and compaction.

The inventory is a spreadsheet (~1000 rows) with one row per firewall. For the AI we
don't send every row; we collapse identical (model, family, version) rows into
"combos" with a stable integer id and send only those. The analyzer expands the AI's
chosen combo ids back into the full list of firewall names.
"""

from __future__ import annotations

import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required. Install with: pip install -r requirements.txt")


def load_inventory(path, sheet=None):
    """Read the inventory into a list of {name, model, family, version} dicts.

    Column headers are matched fuzzily so minor variations (e.g. the leading space in
    " IOS version") still work. Firewalltype ASAv is folded into ASA.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def col(*names):
        for i, h in enumerate(header):
            if any(n in h for n in names):
                return i
        return None

    ci_name = col("firewallname", "name")
    ci_model = col("model")
    ci_type = col("firewalltype", "type")
    ci_ver = col("version", "ios")
    if ci_name is None or ci_type is None or ci_ver is None:
        sys.exit(
            "Inventory is missing required columns. Expected a name column, a "
            "firewall-type column, and an IOS-version column.\n"
            f"Found headers: {rows[0]}"
        )

    inv = []
    for r in rows[1:]:
        if not r or ci_name >= len(r) or r[ci_name] is None:
            continue
        raw_type = str(r[ci_type]).strip().upper() if _cell(r, ci_type) else ""
        family = "ASA" if raw_type.startswith("ASA") else raw_type  # ASAv -> ASA
        inv.append(
            {
                "name": str(r[ci_name]).strip(),
                "model": str(_cell(r, ci_model)).strip() if _cell(r, ci_model) else "",
                "family": family,
                "version": str(_cell(r, ci_ver)).strip() if _cell(r, ci_ver) else "",
            }
        )
    return inv


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def build_combos(inventory):
    """Collapse the inventory into distinct (model, family, version) combos.

    Returns (combos, combo_map):
      combos    : list of {id, model, family, version, count} ordered by id
      combo_map : {id -> [firewall_name, ...]}
    """
    order = []
    index = {}
    combo_map = {}
    for fw in inventory:
        key = (fw["model"], fw["family"], fw["version"])
        if key not in index:
            cid = len(order)
            index[key] = cid
            order.append(key)
            combo_map[cid] = []
        combo_map[index[key]].append(fw["name"])

    combos = [
        {
            "id": cid,
            "model": model,
            "family": family,
            "version": version,
            "count": len(combo_map[cid]),
        }
        for cid, (model, family, version) in enumerate(order)
    ]
    return combos, combo_map


def expand_combo_ids(combo_ids, combo_map):
    """Turn a list of combo ids into a sorted, de-duplicated list of firewall names."""
    names = set()
    for cid in combo_ids or []:
        try:
            cid = int(cid)
        except (TypeError, ValueError):
            continue
        names.update(combo_map.get(cid, []))
    return sorted(names)
